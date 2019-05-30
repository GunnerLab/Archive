"""
Automate mcce calculations for testing a new mcce version.
Description
-----------
This code is used to compare mcce new runs to proteins in this directory (good_proteins)
good_proteins dir is needed to run this code (/home/salah/benchmark_proteins/good_proteins/)


Examples
--------
First run: python run_mcce_test.py -i proteins_list.txt (this creates new directory with the same proteins in good_proteins for example "outputs1_7-17-2016")
Second run: python run_test.py -d outputs1_7-17-2016

"""
#!/usr/bin/python
import os
import datetime
import time
from datetime import timedelta
from shutil import copyfile # This is needed to copy files
from subprocess import call # This is needed to submit jobs
from argparse import ArgumentParser
from shutil import move
from tempfile import mkstemp
from os import remove, close
import numpy as np
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
import xlrd
import xlwt






greenFill = PatternFill(start_color='FF4bf11e',
                   end_color='FF4bf11e',
                   fill_type='solid')


def checkHead3(dir_of_head3):
	# Reading original head3.lst file
	
	protein_name = dir_of_head3.split('/')
	
	head3_original = '/home/salah/benchmark_proteins/good_proteins/'+protein_name[1]+'/head3.lst'
	
	
	desolv_original = []
	epol_original = []
	desolv_list = []
	epol_list = []
	
	lines_ = open(head3_original).readlines()
	
	for line in lines_:
		line = line.split()
		if len(line) != 16:
			return 'In '+dir_of_head3+' head3.lst does not have 15 columns'
		else:
			desolv_original.append(line[13])
			epol_original.append(line[12])
	# Reading new head3.lst file
	desolv = []
	epol = []
	if not os.path.exists(dir_of_head3+'/head3.lst'):
		return 'head3.lst does not exist in '+dir_of_head3
	else:
		head3_lst = dir_of_head3+'/head3.lst'
		lines_ = open(head3_lst).readlines()
		for line in lines_:
			line = line.split()
			if len(line) != 16:
				return 'In '+dir_of_head3+' head3.lst does not have 15 columns'
			else:
				desolv.append(line[13])
				epol.append(line[12])

	# Comparing desolv in both
	
	if len(desolv_original) != len(desolv):
		print 'head3.lst do not have the same length'
	else:
		A = desolv_original
		A.remove('dsolv')
		A = map(float, A)
		A = np.asarray(A)
		B = desolv
		B.remove('dsolv')
		B = map(float, B)
		B = np.asarray(B)
		desolv_diff = A - B
		
		
		sum_of_absolute_differences_desolv = np.sum(np.abs(A - B))
		sum_of_squared_differences_desolv =  np.sum(np.square(A - B))
		correlation_desolv = np.corrcoef(np.array((A, B)))[0, 1]
		standard_deviation_desolv = np.std(A-B)
		
		desolv_list.append(max(desolv_diff))
		desolv_list.append(min(desolv_diff))
		desolv_list.append(sum_of_absolute_differences_desolv)
		desolv_list.append(sum_of_squared_differences_desolv)
		desolv_list.append(correlation_desolv)
		desolv_list.append(standard_deviation_desolv)
		
	
	
	# Comparing epol in both
	
	if len(epol_original) != len(epol):
		print 'head3.lst do not have the same length'
	else:
		A = epol_original
		A.remove('epol')
		A = map(float, A)
		A = np.asarray(A)
		B = epol
		B.remove('epol')
		B = map(float, B)
		B = np.asarray(B)
		epol_diff = A - B
		
		
		sum_of_absolute_differences_epol = np.sum(np.abs(A - B))
		sum_of_squared_differences_epol =  np.sum(np.square(A - B))
		correlation_epol = np.corrcoef(np.array((A, B)))[0, 1]
		standard_deviation_epol = np.std(A-B)
		
		epol_list.append(max(epol_diff))
		epol_list.append(min(epol_diff))
		epol_list.append(sum_of_absolute_differences_epol)
		epol_list.append(sum_of_squared_differences_epol)
		epol_list.append(correlation_epol)
		epol_list.append(standard_deviation_epol)
		
	#desolv_together = np.column_stack((desolv_original,desolv))
	#epol_together = np.column_stack((epol_original,epol))
	
	return desolv_original,desolv, epol_original,epol, desolv_list,epol_list

def checkpKout(dir_of_pKout):
	
	protein_name = dir_of_pKout.split('/')
	pKout_original = '/home/salah/benchmark_proteins/good_proteins/'+protein_name[1]+'/pK.out'
	lines_ = open(pKout_original).readlines()
	
	# Reading original pKout
	pKout_original_list = []
	for line in lines_:
		line = line.split()
		#print 'pKout_org ' + str(len(line))
		if len(line) == 15:
			pKout_original_list.append(line[1])
		else:
			continue
		
	# Reading new head3.lst file
	pKout_new_list = []
	if not os.path.exists(dir_of_pKout+'/pK.out'):
		return 'pK.out does not exist in '+dir_of_pKout
	else:
		pK_out = dir_of_pKout+'/pK.out'
		lines_ = open(pK_out).readlines()
		for line in lines_:
			line = line.split()
			#print 'pKout_new ' + str(len(line))
			if len(line) == 15:
				pKout_new_list.append(line[1])
			else:
				continue
	
	if len(pKout_new_list) != len(pKout_original_list):
		pKout_original_list = [0] * 100
		pKout_new_list = [0] * 100
		return pKout_original_list, pKout_new_list 
	else:
		return pKout_original_list, pKout_new_list
		
				
def desolv_fun(desolv_original,desolv,desolv_list,ws3):
	# Write header
	j = 0
	cell_header1 = 'A1' 
	ws3[cell_header1] = 'desolv original'
	cell_header2 = 'B1' 
	ws3[cell_header2] = 'desolv new'
	
	for desolv_ori in desolv_original:
		cell_1 = 'A'+str(j+2)
		cell_2 = 'B'+str(j+2)
		ws3[cell_1] = float(desolv_ori)
		ws3[cell_2] = float(desolv[j])
		j += 1
	
	diff_desolv = list(np.array([float(j) for j in desolv_original])- np.array([float(j) for j in desolv]))
	if max(diff_desolv) < 0.1:
		cell_header1 = 'C1' 
		ws3[cell_header1] = 'PASS'
		ws3['C1'].fill = greenFill
		
		
def epol_fun(epol_original,epol,epol_list,ws3):
	# Write header
	j = 0
	cell_header1 = 'D1' 
	ws3[cell_header1] = 'epol original'
	cell_header2 = 'E1' 
	ws3[cell_header2] = 'epol new'
	for epol_ori in epol_original:
		#sheet.write(j+2,0,str(statN))
		cell_1 = 'D'+str(j+2)
		cell_2 = 'E'+str(j+2)
		ws3[cell_1] = float(epol_ori)
		ws3[cell_2] = float(epol[j])
		j += 1
	diff_epol = list(np.array([float(j) for j in epol_original])- np.array([float(j) for j in epol]))
	if max(diff_epol) < 0.1:
		cell_header1 = 'F1' 
		ws3[cell_header1] = 'PASS'
		ws3['F1'].fill = greenFill

def pKout_fun(pKout_original_list, pKout_new_list,ws3):
	# Write header
	j = 0
	#print pKout_original_list, pKout_new_list
	cell_header1 = 'G1' 
	ws3[cell_header1] = 'pK original'
	cell_header2 = 'H1' 
	ws3[cell_header2] = 'pK new'
	for pK_ori in pKout_original_list[1:]:
		#sheet.write(j+2,0,str(statN))
		cell_1 = 'G'+str(j+2)
		cell_2 = 'H'+str(j+2)
		ws3[cell_1] = float(pK_ori)
		ws3[cell_2] = float(pKout_new_list[j+1])
		j += 1
	pKout_epol = list(np.array([float(j) for j in pKout_original_list[1:]])- np.array([float(j) for j in pKout_new_list[1:]]))
	if max(pKout_epol) < 0.1:
		cell_header1 = 'I1' 
		ws3[cell_header1] = 'PASS (The maximum difference is less than 0.1)'
		ws3['I1'].fill = greenFill
		
def testing(input_directory):
	i = 0
	wb = openpyxl.Workbook()
	dirs_in_new_dir = [d for d in os.listdir(input_directory) if os.path.isdir(os.path.join(input_directory, d))]
	
	for dir in dirs_in_new_dir:
		print 'Testing '+dir+' head3.lst'
		#time.sleep(1.0)
		
		desolv_original,desolv, epol_original,epol, desolv_list,epol_list = checkHead3(input_directory+dir)
		
		diff_desolv = list(np.array([float(j) for j in desolv_original])- np.array([float(j) for j in desolv]))
		diff_epol = list(np.array([float(j) for j in epol_original])- np.array([float(j) for j in epol]))
		
		if max(diff_desolv) < 0.1:
			print '    desolv energy pass'
			
		if max(diff_epol) < 0.1:
			print '    epol energy pass'
		
		
		pKout_original_list, pKout_new_list = checkpKout(input_directory+dir)
		
		ws3 = wb.create_sheet(index=i, title=dir)
		
		# Write desolv energy to excel
		desolv_fun(desolv_original,desolv,desolv_list,ws3)

		# Write epol energy to excel
		epol_fun(epol_original,epol,epol_list,ws3)
		
		# Write pK.out to excel
		pKout_fun(pKout_original_list, pKout_new_list,ws3)
		
		
		i += 1
	
	wb.save('data.xlsx')

def parse_args():
    """Parse the command line arguments and perform some validation on the
    arguments
    Returns
    -------
    args : argparse.Namespace
        The namespace containing the arguments
    """
    parser = ArgumentParser(description='''Compare one directory to /home/salah/benchmark_proteins/good_proteins/''')
    required = parser.add_argument_group('required argument')
    required.add_argument('-d', '--input_directory', required=True, type=str,
                          help='''path to the directory containing input
                          new mcce jobs.''')
    args = parser.parse_args()
    return args

def main():
	# 
    args = parse_args()
    testing(args.input_directory)
	
	
# Using entry point approach for future conda packaging
def entry_point():
    main()
			
if __name__ == '__main__':
    entry_point()