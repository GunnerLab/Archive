"""
Getting data from mcce runs

Description
-----------
This code is used to get data from mcce runs.  If the directory has only one mcce run, then the code will 
get the data from that run.  If you have sub directories in this directory, the code will also get data from there.


Examples
--------
python getdata.py -d dir_name

"""
import sys
import os
import time
import os.path
from tempfile import mkstemp
from shutil import move
from os import remove, close
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils.indexed_list import IndexedList
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Style
from argparse import ArgumentParser
from argparse import RawTextHelpFormatter

greenFill = PatternFill(start_color='FF4bf11e',
                   end_color='FF4bf11e',
                   fill_type='solid')

				   
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

my_style = Style(border=thin_border)
							
inhibitor_list = [
	'032',
	'0WN',
	'1N1',
	'40L',
	'AQ4',
	'B49',
	'BOS',
	'EMH',
	'FMM',
	'KIM',
	'LQQ',
	'NIL',
	'STI',
	'VGH',
	'YY3',
	'0LI',
	'1E8',
	'276',
	'4MK',
	'AXI',
	'BAX',
	'CAB',
	'EUI',
	'IRE',
	'LEV',
	'MI1',
	'P06',
	'RXT',
	'TRA',
	'XIN',
	'ZD6',
	'DB8',
]

charge_dic_1 = {'0':0,
				'+':1,
				'-':-1,
				'D':0,
}

charge_dic_2 = {'1':1,
			  '2':1,
			  '3':1,
			  '4':1,
			  '5':1,
			  '6':1,
			  '7':1,
			  'a':2,
			  'b':2,
			  'c':2,
			  'd':2,
			  'A':3,
			  'B':3,
			  'M':0,
}

def setting_up(input_dir):
	dir_list_1 = []
	dir_list_2 = []
	dirs = [d for d in os.listdir(input_dir) if os.path.isdir(os.path.join(input_dir, d))]
	pdb_dic = {}
	for dir1 in dirs:
		sub_dirs = [d for d in os.listdir(input_dir+dir1) if os.path.isdir(os.path.join(input_dir+dir1, d))]
		dir_list_1.append(dir1)
		for dir2 in sub_dirs:
			dir_list_2.append(dir2)
			pdb_dic[dir2] = dir1
	return dir_list_1, dir_list_2,pdb_dic

def conclustion(wb):
	ws3 = wb.create_sheet(index=0, title='Conclustion') 
	cell_header1 = 'A1' 
	ws3[cell_header1] = 'Conclusions:'

def fort38(input_dir,dir_list_2,pdb_dic,wb):
	
	numberoflines = 0
	for dir2 in dir_list_2:
		sheetName = dir2.split('_')
		if '_together' in dir2:
			fort38_together = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_together).readlines()
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'A5' 
			ws3[cell_header1] = 'fort38_together'
			i = 7
			count = 0
			for line in reversed(lines_[1:]):
				lineSplit = line.split()
				if lineSplit[0][:3] in inhibitor_list:
					count += 1
				cell_res_name = lineSplit[0]
				sum_occupancy = 0
				for occupancy in lineSplit[1:]:
					sum_occupancy += float(occupancy)
				ave_occupancy = sum_occupancy/5
				cellheader = 'A'+str(i)
				ws3[cellheader] = cell_res_name
				cellheader = 'B'+str(i)
				ws3[cellheader] = ave_occupancy
				i += 1
		elif '_proteinalone' in dir2:
			fort38_proteinalone = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_proteinalone).readlines()
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'C5' 
			ws3[cell_header1] = 'fort38_proteinalone'
			i = 7+count
			for line in reversed(lines_[1:]):
				lineSplit = line.split()
				cell_res_name = lineSplit[0]
				sum_occupancy = 0
				for occupancy in lineSplit[1:]:
					sum_occupancy += float(occupancy)
				ave_occupancy = sum_occupancy/5
				cellheader = 'C'+str(i)
				ws3[cellheader] = cell_res_name
				cellheader = 'D'+str(i)
				ws3[cellheader] = ave_occupancy
				i += 1
		elif '_inhibitoralone' in dir2:
			fort38_inhibitoralone = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_inhibitoralone).readlines()
			numberoflines = len(lines_) # Needed to line up the important part of the protein (the inhibitor)
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'E5' 
			ws3[cell_header1] = 'fort38_inhibitoralone'
			i = 7
			for line in reversed(lines_[1:]):
				lineSplit = line.split()
				cell_res_name = lineSplit[0]
				sum_occupancy = 0
				for occupancy in lineSplit[1:]:
					sum_occupancy += float(occupancy)
				ave_occupancy = sum_occupancy/5
				cellheader = 'E'+str(i)
				ws3[cellheader] = cell_res_name
				cellheader = 'F'+str(i)
				ws3[cellheader] = ave_occupancy
				i += 1
			
def sum_crg(input_dir,dir_list_2,pdb_dic,wb):
	for dir2 in dir_list_2:
		sheetName = dir2.split('_')
		#print input_dir,pdb_dic[dir2],dir2
		if '_together' in dir2:
			sum_crg_together = input_dir+pdb_dic[dir2]+'/'+dir2+'/sum_crg.out'
			lines_ = open(sum_crg_together).readlines()
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'H5' 
			ws3[cell_header1] = 'sum_crg_together'
			i = 7
			for line in reversed(lines_):
				if len(line.strip()) != 0:   # Make sure the line is not empty 
					lineSplit = line.split()
					cell_res_name = lineSplit[0]+lineSplit[1]
					sum_charge = 0
					for occupancy in lineSplit[2:]:
						sum_charge += float(occupancy)
					ave_charge = sum_charge/5
					cellheader = 'H'+str(i)
					ws3[cellheader] = cell_res_name
					cellheader = 'I'+str(i)
					ws3[cellheader] = ave_charge
					if 'crg' in lineSplit[1] or 'Elec' in lineSplit[1] or 'Prot' in lineSplit[1]:
						cellheader = 'H'+str(i)
						ws3[cellheader].style = my_style
						cellheader = 'I'+str(i)
						ws3[cellheader].style = my_style
					i += 1
					
					
		elif '_proteinalone' in dir2:
			sum_crg_proteinalone = input_dir+pdb_dic[dir2]+'/'+dir2+'/sum_crg.out'
			lines_ = open(sum_crg_proteinalone).readlines()
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'J5' 
			ws3[cell_header1] = 'sum_crg_proteinalone'
			i = 7
			for line in reversed(lines_):
				if len(line.strip()) != 0:
					lineSplit = line.split()
					cell_res_name = lineSplit[0]+lineSplit[1]
					sum_charge = 0
					for occupancy in lineSplit[2:]:
						sum_charge += float(occupancy)
					ave_charge = sum_charge/5
					cellheader = 'J'+str(i)
					ws3[cellheader] = cell_res_name
					cellheader = 'K'+str(i)
					ws3[cellheader] = ave_charge
					if 'crg' in lineSplit[1] or 'Elec' in lineSplit[1] or 'Prot' in lineSplit[1]:
						cellheader = 'J'+str(i)
						ws3[cellheader].style = my_style
						cellheader = 'K'+str(i)
						ws3[cellheader].style = my_style
					i += 1
					
def sum_crg_lig(input_dir,dir_list_2,pdb_dic,wb):
	for dir2 in dir_list_2:
		sheetName = dir2.split('_')
		if '_together' in dir2:
			fort38_together = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_together).readlines()
			ws3 = wb.get_sheet_by_name(sheetName[0]+'_'+pdb_dic[dir2])
			cell_header1 = 'L5' 
			ws3[cell_header1] = 'sum_crg_inhibitoralone'
			i = 10
			total_charge = 0
			for line in reversed(lines_[1:]):
				lineSplit = line.split()
				if lineSplit[0][:3] in inhibitor_list:
					sum_occupancy = 0
					for occupancy in lineSplit[1:]:
						sum_occupancy += float(occupancy)
					ave_occupancy = sum_occupancy/5
					cell_header1 = 'L'+str(i) 
					ws3[cell_header1] = lineSplit[0]
					cell_header1 = 'M'+str(i) 
					ws3[cell_header1] = ave_occupancy * charge_dic_1[lineSplit[0][3]]*charge_dic_2[lineSplit[0][4]]
					total_charge += ave_occupancy * charge_dic_1[lineSplit[0][3]]*charge_dic_2[lineSplit[0][4]]
					i += 1
					
			cell_header1 = 'L7' 
			ws3[cell_header1] = 'Netcrg'
			ws3['L7'].style = my_style
			cell_header1 = 'M7' 
			ws3[cell_header1] = total_charge
			ws3['M7'].style = my_style
			cell_header1 = 'L8' 
			ws3[cell_header1] = 'TotElec'
			ws3['L8'].style = my_style
			cell_header1 = 'M8' 
			ws3[cell_header1] = 0
			ws3['M8'].style = my_style
			cell_header1 = 'L9' 
			ws3[cell_header1] = 'TotProtn'
			ws3['L9'].style = my_style
			cell_header1 = 'M9' 
			ws3[cell_header1] = total_charge
			ws3['M9'].style = my_style
			
def delta_crg_in_protein(input_dir,dir_list_2,pdb_dic,wb):
	ws3 = wb.get_sheet_by_name('delta crg')
	i = 3
	for dir2 in dir_list_2:
		if '_together' in dir2:
			pdbName = dir2.split('_')
			cell_header1 = 'B'+str(i) 
			ws3[cell_header1] = pdb_dic[dir2]
			cell_header1 = 'C'+str(i) 
			ws3[cell_header1] = pdbName[0]
			i += 1
	j = 3
	k = 3
	ii = 3
	jj = 3
	record_index = 3
	for dir2 in dir_list_2:
		sheetName = dir2.split('_')
		sum_crg = 0
		ave_crg = 0
		count = 0
		occ_inhibitor_list_together = []
		conf_inhibitor_list_together = []
		occ_inhibitor_list = []
		conf_inhibitor_list = []
		if '_together' in dir2:
			cell_header1 = 'M'+str(record_index)
			ws3[cell_header1] = input_dir+pdb_dic[dir2]+'/'+sheetName[0]
			cell_header1 = 'N'+str(record_index)
			ws3[cell_header1] = ' '
			record_index += 1
			sum_crg_together = input_dir+pdb_dic[dir2]+'/'+dir2+'/sum_crg.out'
			lines_ = open(sum_crg_together).readlines()
			for line in lines_:
				if len(line.strip()) != 0:   # Make sure the line is not empty 
					lineSplit = line.split()
					if 'crg' in lineSplit:
						#print lineSplit
						cell_header1 = 'D'+str(j)
						for item in lineSplit[2:]:
							sum_crg += float(item)
							count += 1
						ave_crg = sum_crg/count
						ws3[cell_header1] = round(ave_crg, 2)
						j += 1
			# get sum charge from fort.38
			fort38_together = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_together).readlines()
			total_crg = 0
			for line in lines_:
				lineSplit = line.split()
				if lineSplit[0][:3] in inhibitor_list:
					#print dir2, lineSplit
					chargeinfort38 = float(lineSplit[1]) * charge_dic_1[lineSplit[0][3]]*charge_dic_2[lineSplit[0][4]]
					total_crg = total_crg + chargeinfort38
					occ_inhibitor_list_together.append(float(lineSplit[1]))
					conf_inhibitor_list_together.append(lineSplit[0])
			cell_header1 = 'F'+str(ii)
			ws3[cell_header1] = round(total_crg, 2)
			cell_header1 = 'I'+str(ii)
			ws3[cell_header1] =  conf_inhibitor_list_together[occ_inhibitor_list_together.index(max(occ_inhibitor_list_together))]
			ii += 1
		elif '_proteinalone' in dir2:
			sum_crg_proteinalone = input_dir+pdb_dic[dir2]+'/'+dir2+'/sum_crg.out'
			lines_ = open(sum_crg_proteinalone).readlines()
			for line in lines_:
				if len(line.strip()) != 0:   # Make sure the line is not empty 
					lineSplit = line.split()
					if 'crg' in lineSplit:
						#print lineSplit
						cell_header1 = 'E'+str(k) 
						for item in lineSplit[2:]:
							sum_crg += float(item)
							count += 1
						ave_crg = sum_crg/count
						
						ws3[cell_header1] = round(ave_crg, 2)
						k += 1
		elif '_inhibitoralone' in dir2:
			# get sum charge from fort.38
			fort38_proteinalone = input_dir+pdb_dic[dir2]+'/'+dir2+'/fort.38'
			lines_ = open(fort38_proteinalone).readlines()
			total_crg = 0
			for line in lines_:
				lineSplit = line.split()
				if lineSplit[0][:3] in inhibitor_list:
					#print dir2, lineSplit
					chargeinfort38 = float(lineSplit[1]) * charge_dic_1[lineSplit[0][3]]*charge_dic_2[lineSplit[0][4]]
					total_crg = total_crg + chargeinfort38
					occ_inhibitor_list.append(float(lineSplit[1]))
					conf_inhibitor_list.append(lineSplit[0])
			cell_header1 = 'G'+str(jj)
			ws3[cell_header1] = round(total_crg, 2)
			cell_header1 = 'H'+str(jj)
			ws3[cell_header1] =  conf_inhibitor_list[occ_inhibitor_list.index(max(occ_inhibitor_list))] 
			jj += 1
			
		
			
def delta_crg(input_dir,dir_list_2,pdb_dic,wb):
	ws3 = wb.create_sheet(index=1, title='delta crg') 
	cell_header1 = 'B1' 
	ws3[cell_header1] = 'inhibitor'	
	cell_header1 = 'C1'
	ws3[cell_header1] = 'pdb'
	cell_header1 = 'D1'
	ws3[cell_header1] = 'protein charge with inhibitor'
	cell_header1 = 'E1'
	ws3[cell_header1] = 'protein charge apo'
	cell_header1 = 'F1'
	ws3[cell_header1] = 'inhibitor charge with prot'
	cell_header1 = 'G1'
	ws3[cell_header1] = 'inhibitor charge in soln'
	cell_header1 = 'H1'
	ws3[cell_header1] = 'most occupied conformer of inhibitor in solution'
	cell_header1 = 'I1'
	ws3[cell_header1] = 'most occupied conformer of inhibitor in protein'
	
	
	
	cell_header1 = 'J1'
	ws3[cell_header1] = 'delta charge prot'
	cell_header1 = 'K1'
	ws3[cell_header1] = 'delta charge inhibitor'
	cell_header1 = 'L1'
	ws3[cell_header1] = 'delta conformation inhibitor'
	cell_header1 = 'M1'
	ws3[cell_header1] = 'record'
	cell_header1 = 'N1'
	ws3[cell_header1] = 'Comments'
	
	delta_crg_in_protein(input_dir,dir_list_2,pdb_dic,wb)	
		
def seeting_up_excel(dir_list_2, pdb_dic,input_dir):
	wb = openpyxl.Workbook()
	
	i = 2
	for pdb_dir in dir_list_2:
		if '_together' in pdb_dir:
			sheetName = pdb_dir.split('_')
			ws3 = wb.create_sheet(index=i, title=sheetName[0]+'_'+pdb_dic[pdb_dir]) 
			
			# Sheet header
			location_of_prot_dir = input_dir+sheetName[0]+pdb_dic[pdb_dir]
			cell_header1 = 'A1' 
			ws3[cell_header1] = location_of_prot_dir
			i += 1	
			
			
	# Doing fort.38 
	fort38(input_dir,dir_list_2,pdb_dic,wb)
	
	# Doing sum_crg.out
	sum_crg(input_dir,dir_list_2,pdb_dic,wb)
	
	# Doing sum_crg_lig
	sum_crg_lig(input_dir,dir_list_2,pdb_dic,wb)
	
	# Conclustion
	conclustion(wb)
	
	# Doing delta charge 
	delta_crg(input_dir,dir_list_2,pdb_dic,wb)
	
	
	
	
	wb.save('data.xlsx')
		
def parse_args():
	"""Parse the command line arguments and perform some validation on the
	arguments
	Returns
	-------
	args : argparse.Namespace
	The namespace containing the arguments
	"""
	parser = ArgumentParser(description='''Run MCCE on multiple PDB files''',formatter_class=RawTextHelpFormatter)
	required = parser.add_argument_group('required argument')
	required.add_argument('-d', '--str_1', required=True, type=str,
                          help='''Enter the directory where the mcce runs are.\nFor example, python getdata.py -d dir_name''')
	args = parser.parse_args()
	return args
	
def main():
	args = parse_args()
	input_dir = args.str_1 
	
	# Setting up the directories
	dir_list_1, dir_list_2, pdb_dic = setting_up(input_dir)
	
	# Setting up excel 
	seeting_up_excel(dir_list_2,pdb_dic,input_dir)
			
def entry_point():
    main()
			
if __name__ == '__main__':
    entry_point()
	
